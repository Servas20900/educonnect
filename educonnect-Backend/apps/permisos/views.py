from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser
from django.utils import timezone
from apps.databaseModels.models import AuthUsuario, AuthRol, AuthPermiso, AuthUsuarioRol, AuthRolPermiso
from .models import ConfiguracionSistema
from .defaults import DEFAULT_CONFIGURATION_MAP
from .serializers import (
    UsuarioListSerializer, UsuarioDetailSerializer, UsuarioUpdateSerializer,
    RolSerializer, RolUpdateSerializer, PermisoSerializer, ModuloSerializer
)


class UsuarioViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de usuarios
    """
    queryset = AuthUsuario.objects.all().order_by('-fecha_registro')
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action == 'list':
            return UsuarioListSerializer
        elif self.action in ['update', 'partial_update']:
            return UsuarioUpdateSerializer
        return UsuarioDetailSerializer
    
    @action(detail=True, methods=['post'])
    def toggle_active(self, request, pk=None):
        """Activa o desactiva un usuario"""
        usuario = self.get_object()
        usuario.is_active = not usuario.is_active
        usuario.save()
        return Response({
            'message': f'Usuario {"activado" if usuario.is_active else "desactivado"} exitosamente',
            'is_active': usuario.is_active
        })
    
    @action(detail=True, methods=['post'])
    def assign_role(self, request, pk=None):
        """Agrega un rol al usuario sin borrar los anteriores"""
        usuario = self.get_object()
        rol_id = request.data.get('rol_id')

        if not rol_id:
            return Response(
                {'error': 'rol_id es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            rol = AuthRol.objects.get(id=rol_id, activo=True)
        except AuthRol.DoesNotExist:
            return Response(
                {'error': 'Rol no encontrado o inactivo'},
                status=status.HTTP_404_NOT_FOUND
            )

        _, created = AuthUsuarioRol.objects.get_or_create(
            usuario=usuario,
            rol=rol,
            defaults={'fecha_asignacion': timezone.now()}
        )

        if not created:
            return Response(
                {'message': f'El usuario ya tiene el rol {rol.nombre}'}
            )

        return Response(
            {'message': f'Rol {rol.nombre} agregado exitosamente'}
        )

    @action(detail=True, methods=['post'])
    def remove_role(self, request, pk=None):
        """Quita un rol especifico del usuario"""
        usuario = self.get_object()
        rol_id = request.data.get('rol_id')

        if not rol_id:
            return Response(
                {'error': 'rol_id es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        deleted, _ = AuthUsuarioRol.objects.filter(
            usuario=usuario,
            rol_id=rol_id
        ).delete()

        if deleted == 0:
            return Response(
                {'error': 'El usuario no tiene ese rol'},
                status=status.HTTP_404_NOT_FOUND
            )

        return Response(
            {'message': 'Rol removido exitosamente'}
        )

    @action(detail=False, methods=['post'],
            parser_classes=[MultiPartParser])
    def import_bulk(self, request):
        """
        Importa usuarios desde un archivo CSV o Excel.
        El archivo debe tener columnas:
        username, email, nombre, primer_apellido,
        segundo_apellido (opcional), fecha_nacimiento
        (YYYY-MM-DD), genero
        Devuelve { creados: N, errores: [...] }
        """
        archivo = request.FILES.get('file')
        if not archivo:
            return Response(
                {'error': 'No se proporciono ningun archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )

        nombre_archivo = archivo.name.lower()
        creados = 0
        errores = []

        try:
            if nombre_archivo.endswith('.csv'):
                import csv, io
                contenido = archivo.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(contenido))
                filas = list(reader)
            elif nombre_archivo.endswith(('.xlsx', '.xls')):
                import openpyxl
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                headers = [str(cell.value).strip() for cell in ws[1]]
                filas = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    filas.append(dict(zip(headers, row)))
            else:
                return Response(
                    {'error': 'Formato no soportado. Use CSV o XLSX.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            from apps.databaseModels.serializers import RegistroSerializer
            for i, fila in enumerate(filas, start=2):
                try:
                    data = {
                        'username': str(fila.get('username', '') or '').strip(),
                        'email': str(fila.get('email', '') or '').strip(),
                        'password': str(fila.get('username', '') or '').strip(),
                        'nombre': str(fila.get('nombre', '') or '').strip(),
                        'primer_apellido': str(fila.get('primer_apellido', '') or '').strip(),
                        'segundo_apellido': str(fila.get('segundo_apellido', '') or '').strip(),
                        'fecha_nacimiento': str(fila.get('fecha_nacimiento', '') or '').strip(),
                        'genero': str(fila.get('genero', '') or 'No especificado').strip(),
                    }
                    if not data['username'] or not data['email']:
                        errores.append(
                            f'Fila {i}: username y email son requeridos'
                        )
                        continue

                    serializer = RegistroSerializer(data=data)
                    if serializer.is_valid():
                        serializer.save()
                        creados += 1
                    else:
                        errores.append(
                            f'Fila {i} ({data["email"]}): '
                            f'{serializer.errors}'
                        )
                except Exception as e:
                    errores.append(f'Fila {i}: {str(e)}')

        except Exception as e:
            return Response(
                {'error': f'Error al procesar el archivo: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        return Response({'creados': creados, 'errores': errores})


class RolViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de roles
    """
    queryset = AuthRol.objects.all().order_by('nombre')
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return RolUpdateSerializer
        return RolSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        # Filtrar solo roles activos si se solicita
        if self.request.query_params.get('activos_only') == 'true':
            queryset = queryset.filter(activo=True)
        return queryset
    
    @action(detail=True, methods=['post'])
    def toggle_active(self, request, pk=None):
        """Activa o desactiva un rol"""
        rol = self.get_object()
        rol.activo = not rol.activo
        rol.fecha_modificacion = timezone.now()
        rol.save()
        return Response({
            'message': f'Rol {"activado" if rol.activo else "desactivado"} exitosamente',
            'activo': rol.activo
        })
    
    @action(detail=True, methods=['post'])
    def update_permissions(self, request, pk=None):
        """Actualiza los permisos de un rol"""
        rol = self.get_object()
        permisos_ids = request.data.get('permisos_ids', [])
        
        # Eliminar permisos actuales
        AuthRolPermiso.objects.filter(rol=rol).delete()
        
        # Asignar nuevos permisos
        for permiso_id in permisos_ids:
            try:
                permiso = AuthPermiso.objects.get(id=permiso_id, activo=True)
                AuthRolPermiso.objects.create(
                    rol=rol,
                    permiso=permiso,
                    fecha_asignacion=timezone.now()
                )
            except AuthPermiso.DoesNotExist:
                continue
        
        return Response({'message': 'Permisos actualizados exitosamente'})


class PermisoViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet para listar permisos (solo lectura)
    """
    queryset = AuthPermiso.objects.filter(activo=True).order_by('modulo', 'nombre')
    serializer_class = PermisoSerializer
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def by_module(self, request):
        """Agrupa permisos por módulo"""
        permisos = self.get_queryset()
        modulos_dict = {}
        
        for permiso in permisos:
            if permiso.modulo not in modulos_dict:
                modulos_dict[permiso.modulo] = []
            modulos_dict[permiso.modulo].append({
                'id': permiso.id,
                'nombre': permiso.nombre,
                'descripcion': permiso.descripcion,
                'accion': permiso.accion
            })
        
        return Response(modulos_dict)


class ModuloViewSet(viewsets.ViewSet):
    """
    ViewSet para listar los módulos del sistema (definidos estáticamente)
    """
    permission_classes = [IsAuthenticated]

    @staticmethod
    def _config_value_or_default(clave):
        default_value = DEFAULT_CONFIGURATION_MAP[clave]['valor']
        config = ConfiguracionSistema.objects.filter(clave=clave, activo=True).first()
        if not config:
            return default_value
        return config.valor
    
    def list(self, request):
        """Lista todos los módulos disponibles en el sistema"""
        modulos = [
            {
                'id': 'admin',
                'nombre': 'Administración',
                'grupo': 'Admin',
                'submodulos': [
                    'Dashboard', 'Circulares', 'Programación', 'Horarios',
                    'Aprobaciones', 'Reportes', 'Usuarios', 'Permisos',
                    'Incapacidades', 'Comites', 'Oficios/Plantillas',
                    'Repositorios', 'Backups', 'Retención'
                ]
            },
            {
                'id': 'docente',
                'nombre': 'Módulo Docente',
                'grupo': 'Docente',
                'submodulos': [
                    'Dashboard', 'Estudiantes', 'Evaluaciones', 'Calificaciones',
                    'Promedios', 'Planeamientos', 'Comunicados', 'Asistencia',
                    'Exportación de Notas', 'Estudiantes en Riesgo'
                ]
            },
            {
                'id': 'comites',
                'nombre': 'Comités',
                'grupo': 'Comites',
                'submodulos': ['Home', 'Crear Actas', 'Agendar Reunion', 'Roles']
            },
            {
                'id': 'auxiliares',
                'nombre': 'Órganos Auxiliares',
                'grupo': 'Auxiliares',
                'submodulos': [
                    'Informes Económicos', 'Reglamentos', 'Informes PAT'
                ]
            },
            {
                'id': 'estudiante',
                'nombre': 'Módulo Estudiante',
                'grupo': 'Estudiante',
                'submodulos': ['Home', 'Notificaciones', 'Circulares y Horarios']
            }
        ]
        
        serializer = ModuloSerializer(modulos, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def por_rol(self, request):
        """Obtiene los módulos permitidos para cada rol"""
        roles = AuthRol.objects.filter(activo=True)
        configuracion = {}
        
        for rol in roles:
            # Obtener módulos únicos de los permisos del rol
            permisos_rol = AuthRolPermiso.objects.filter(rol=rol).select_related('permiso')
            modulos = set(rp.permiso.modulo for rp in permisos_rol if rp.permiso)
            
            configuracion[rol.nombre.lower()] = {
                'modulos': list(modulos),
                'rol_id': rol.id,
                'tipo_rol': rol.tipo_rol
            }
        
        return Response(configuracion)

    @action(detail=False, methods=['get'])
    def bootstrap(self, request):
        """Devuelve configuración inicial para frontend: autorización, navegación y catálogos."""
        role_names = list(
            AuthUsuarioRol.objects.filter(usuario=request.user)
            .select_related('rol')
            .values_list('rol__nombre', flat=True)
        )
        roles = [str(role_name).lower() for role_name in role_names if role_name]

        if request.user.is_superuser and 'administrador' not in roles:
            roles.insert(0, 'administrador')

        current_role = roles[0] if roles else 'usuario'

        route_permissions = self._config_value_or_default('route_permissions')
        route_permissions['horarios'] = ['administrador']
        route_permissions['documentos'] = ['administrador', 'docente']
        route_permissions.pop('docente-incapacidades', None)
        route_permissions.pop('auxiliares-cumplimiento', None)

        navigation = self._config_value_or_default('navigation')
        if isinstance(navigation, dict):
            for group in navigation.get('items', []):
                if group.get('id') == 'comite-group':
                    # Normaliza configuraciones antiguas para que el modulo Comite
                    # solo aparezca a usuarios con rol comite o administrador.
                    group['allowed_roles'] = ['administrador', 'comite']
                if group.get('id') == 'docente':
                    for child in group.get('children', []):
                        if child.get('id') != 'docente-items':
                            continue
                        child['children'] = [
                            item for item in child.get('children', [])
                            if item.get('id') not in {'docente-incapacidades', 'incapacidades'}
                        ]
                if group.get('id') == 'auxiliares':
                    # Garantiza visibilidad del modulo Auxiliares para roles vigentes y legado.
                    group['allowed_roles'] = ['administrador', 'auxiliares', 'auxiliar']
                    for child in group.get('children', []):
                        if child.get('id') != 'auxiliares-items':
                            continue
                        child['children'] = [
                            item for item in child.get('children', [])
                            if item.get('id') != 'reportes-cumplimiento'
                        ]

        for key in ('auxiliares-informes', 'auxiliares-reglamentos'):
            existing_roles = route_permissions.get(key, [])
            normalized_roles = {str(role).strip().lower() for role in existing_roles if role}
            normalized_roles.update({'administrador', 'auxiliares', 'auxiliar'})
            route_permissions[key] = sorted(normalized_roles)

        allowed_permission_keys = [
            key for key, allowed_roles in route_permissions.items()
            if set(allowed_roles).intersection(set(roles))
        ]

        return Response({
            'user': {
                'username': request.user.username,
                'current_role': current_role,
                'roles': roles,
            },
            'navigation': navigation,
            'route_permissions': route_permissions,
            'allowed_permission_keys': allowed_permission_keys,
            'catalogs': self._config_value_or_default('catalogs'),
            'branding': self._config_value_or_default('branding'),
            'public_nav': self._config_value_or_default('public_nav'),
        })
